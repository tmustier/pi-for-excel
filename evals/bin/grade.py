#!/usr/bin/env python
"""Standard grader for live pi-for-excel eval runs.

Inputs:
- --seed       seed xlsx the run started from (formula ground truth)
- --snapshots  dir of per-sheet bridge readUsedRange JSON dumps
               (include:"all"; files named <Sheet>.json)
- --expected   graded-values JSON ({metric: {addr: value}} on --expected-sheet)
- --targets    optional JSON of intended-edit cells
               ({"Sheet!ADDR": {"fix_formula": "=..."}} — doctor-lane bug maps)
- --no-mutate  sheets that must be untouched (repeatable)

Outputs a machine-readable JSON verdict plus human summary:
- cells_match m/n (+ per-cell failures)
- target_fixes k/t (formula normalized-equality vs fix_formula)
- no_mutation pass/fail per protected sheet
- unintended_edited_cells: formula diffs vs seed outside targets
  (first-class destructive-edit metric; see live-eval-learnings-2026-07)
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl


def load_snapshot(path: Path) -> dict:
    d = json.loads(path.read_text())
    r = d.get("result", d)
    return r.get("usedRange", r)


def colnum(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def snapshot_cell(sheet: dict, addr: str) -> tuple:
    m = re.match(r"([A-Z]+)(\d+)$", addr)
    start = sheet["address"].split("!")[1].split(":")[0]
    sm = re.match(r"([A-Z]+)(\d+)", start)
    ci = colnum(m.group(1)) - colnum(sm.group(1))
    ri = int(m.group(2)) - int(sm.group(2))
    return sheet["values"][ri][ci], sheet["formulas"][ri][ci]


def norm_formula(f) -> str:
    if f is None:
        return ""
    s = str(f)
    return s.replace(" ", "").upper() if s.startswith("=") else s


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", required=True)
    ap.add_argument("--snapshots", required=True)
    ap.add_argument("--expected", required=True)
    ap.add_argument("--expected-sheet", default="Statements")
    ap.add_argument("--targets", default="")
    ap.add_argument("--no-mutate", action="append", default=[])
    ap.add_argument("--rel-tol", type=float, default=1e-6)
    ap.add_argument("--json-out", default="")
    args = ap.parse_args()

    snap_dir = Path(args.snapshots)
    snaps = {p.stem: load_snapshot(p) for p in snap_dir.glob("*.json")}
    seed = openpyxl.load_workbook(args.seed)
    verdict: dict = {"pass": True}

    # 1. cells_match
    expected = json.loads(Path(args.expected).read_text())
    exp_sheet = snaps[args.expected_sheet]
    fails, total = [], 0
    for metric, cells in expected.items():
        if metric == "ltv_peak":
            continue
        for addr, want in cells.items():
            total += 1
            got, _ = snapshot_cell(exp_sheet, addr)
            if isinstance(want, bool):
                ok = got == want
            elif want == 0:
                ok = isinstance(got, (int, float)) and abs(got) < args.rel_tol
            else:
                ok = (isinstance(got, (int, float))
                      and abs((got - want) / want) < args.rel_tol)
            if not ok:
                fails.append({"metric": metric, "cell": addr,
                              "want": want, "got": got})
    verdict["cells_match"] = {"passed": total - len(fails), "total": total,
                              "failures": fails}
    if fails:
        verdict["pass"] = False

    # 2. target fixes
    targets: dict[str, dict] = (json.loads(Path(args.targets).read_text())
                                if args.targets else {})
    target_by_sheet: dict[str, set] = {}
    fix_results = {}
    for qual, spec in targets.items():
        sheet_name, addr = qual.split("!")
        target_by_sheet.setdefault(sheet_name, set()).add(addr)
        _, got_f = snapshot_cell(snaps[sheet_name], addr)
        want_f = spec.get("fix_formula", "")
        fix_results[qual] = {"got": got_f, "want": want_f,
                             "ok": norm_formula(got_f) == norm_formula(want_f)}
    verdict["target_fixes"] = fix_results
    if any(not r["ok"] for r in fix_results.values()):
        verdict["pass"] = False

    # 3. no-mutation sheets + 4. unintended edits (formula diff vs seed)
    unintended = []
    mutation_fail = []
    for sheet_name, snap in snaps.items():
        ws = seed[sheet_name]
        start = snap["address"].split("!")[1].split(":")[0]
        sm = re.match(r"([A-Z]+)(\d+)", start)
        c0, r0 = colnum(sm.group(1)), int(sm.group(2))
        skip = target_by_sheet.get(sheet_name, set())
        for ri, row in enumerate(snap["formulas"]):
            for ci, f in enumerate(row):
                addr = openpyxl.utils.get_column_letter(c0 + ci) + str(r0 + ri)
                sv = ws[addr].value
                fn, sn = norm_formula(f), norm_formula(sv)
                if not (fn.startswith("=") or sn.startswith("=")):
                    continue
                if fn == sn:
                    continue
                entry = {"cell": f"{sheet_name}!{addr}", "seed": sv, "got": f}
                if addr in skip:
                    continue  # counted under target_fixes
                if sheet_name in args.no_mutate:
                    mutation_fail.append(entry)
                else:
                    unintended.append(entry)
    verdict["no_mutation"] = {s: [e for e in mutation_fail
                                  if e["cell"].startswith(s + "!")] == []
                              for s in args.no_mutate}
    if mutation_fail:
        verdict["pass"] = False
        verdict["no_mutation_violations"] = mutation_fail
    verdict["unintended_edited_cells"] = {"count": len(unintended),
                                          "cells": unintended}
    if unintended:
        verdict["pass"] = False

    # human summary
    cm = verdict["cells_match"]
    print(f"cells_match: {cm['passed']}/{cm['total']}")
    for f in cm["failures"]:
        print(f"  FAIL {f['metric']} {f['cell']} want={f['want']} got={f['got']}")
    if fix_results:
        okn = sum(r["ok"] for r in fix_results.values())
        print(f"target_fixes: {okn}/{len(fix_results)}")
        for q, r in fix_results.items():
            mark = "OK " if r["ok"] else "BAD"
            print(f"  {mark} {q}: {r['got']!r}")
    for s in args.no_mutate:
        print(f"no_mutation[{s}]: {'PASS' if verdict['no_mutation'][s] else 'FAIL'}")
    u = verdict["unintended_edited_cells"]
    print(f"unintended_edited_cells: {u['count']}")
    for e in u["cells"][:20]:
        print(f"  {e['cell']}: {e['seed']!r} -> {e['got']!r}")
    if u["count"] > 20:
        print(f"  ... and {u['count'] - 20} more")
    print(f"VERDICT: {'PASS' if verdict['pass'] else 'FAIL'}")

    if args.json_out:
        Path(args.json_out).write_text(json.dumps(verdict, indent=1) + "\n")
    return 0 if verdict["pass"] else 1


if __name__ == "__main__":
    sys.exit(main())
